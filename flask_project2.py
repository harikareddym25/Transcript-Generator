from flask import *
from werkzeug.utils import secure_filename
import datetime
import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border,Font,Color,Alignment
import win32com.client
from openpyxl.drawing.image import Image
import pythoncom

app = Flask(__name__)
@app.route("/success/<name>")
def success(name):
    return "Succesfully Generated Mark Sheets!!" % name

def prepare_file(data, subjects_data, names_data, grade_dict, roll):
    wb = Workbook()
    sem_list = sorted(list(data[roll].keys()), key=lambda x: int(x))
    sheet = wb.active
    bd = openpyxl.styles.Side(style='thin', color="000000")
    normal_style=Font(name='century',size=8)
    highlight = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
    sheet.cell(row=6,column=4).value='RollNo:'
    sheet.cell(row=6,column=5).value=roll
    sheet.cell(row=6,column=7).value='Name:'
    sheet.cell(row=6,column=8).value=names_data[roll]
    sheet.cell(row=7,column=4).value='Programme:'
    sheet.cell(row=7,column=5).value='Bachelor of Technology'
    sheet.cell(row=7,column=7).value='Course:'
    if roll[4:6]== "CS":
        sheet.cell(row=7,column=8).value='Computer Science and Engineering '
    if roll[4:6]== "EE":
        sheet.cell(row=7,column=8).value='Electrical Engineering'
    if roll[4:6]== "ME":
        sheet.cell(row=7,column=8).value='Mechanical Engineering'
        
    for sem_no in sem_list:
        sem_credits = 0
        sem_achieved_credits = 0
        if(sem_no=='1'):
            row_start=8
            col_start=1
        elif(sem_no=='2'):
            row_start=8
            col_start=7
        elif(sem_no=='3'):
            row_start=8
            col_start=13
        elif(sem_no=='4'):
            row_start=21
            col_start=1
        elif(sem_no=='5'):
            row_start=21
            col_start=7
        elif(sem_no=='6'):
            row_start=21
            col_start=13
        elif(sem_no=='7'):
            row_start=34
            col_start=1
        elif(sem_no=='8'):
            row_start=34
            col_start=7
        fieldnames = [
            "Subject No.",
            "Subject Name",
            "L-T-P",
            "Credit",
            "Grade",
        ]
        sheet.cell(row=row_start,column=col_start).value='Semester'+sem_no
        for i, fieldname in enumerate(fieldnames, start=col_start):
            sheet.cell(row=row_start+1, column=i).value = fieldname
            sheet.cell(row=row_start+1, column=i).font = normal_style
            sheet.cell(row=row_start+1, column=i).border=highlight
        for row_no, sub in enumerate(data[roll][sem_no], start=row_start+1):
            sheet_row = [
                #row_no,
                sub["code"],
                subjects_data[sub["code"]]["subname"],
                subjects_data[sub["code"]]["ltp"],
                sub["credit"],
                #sub["subtype"],
                sub["grade"],
            ]
            for col_no, val in enumerate(sheet_row, start=col_start):
                sheet.cell(row=row_no + 1, column=col_no).value = val
                sheet.cell(row=row_no + 1, column=col_no).font = normal_style
                sheet.cell(row=row_no + 1, column=col_no).border = highlight
            sem_credits += sub["credit"]
            sem_achieved_credits += grade_dict[sub["grade"].strip()] * sub["credit"]
            for i in 'ACDEFGIJKLMOPQ':
                sheet.column_dimensions[i].width=8
            for i in range(1,50):
                sheet.row_dimensions[i].height=12
            for i in 'BHN':
                sheet.column_dimensions[i].width=26
        
        img=openpyxl.drawing.image.Image('IITP_logo.JPEG')
        img.height=60
        img.width=1450
        sheet.add_image(img)
        sheet.cell(row=49,column=1).value='Date generated:'#
        current_time=datetime.datetime.now()
        sheet.cell(row=49,column=2).value=current_time
        sheet.cell(row=49,column=15).value='Assistant Registrar(Acadamic)'#
        img1=openpyxl.drawing.image.Image('seal.JPEG')
        img1.height=100
        img1.width=200
        sheet.add_image(img1,'H46')

    location="output"
    if not os.path.exists(location):
        os.makedirs(location)
    file_name=roll + ".xlsx"
    file_path=os.path.join(location,file_name)
    wb.save(file_path)

    
data = {}
with open("grades.csv", "r") as f:
    reader = csv.DictReader(f)
    grades_data = [dict(row) for row in reader]
    for i, row in enumerate(grades_data):
        grades_data[i]["Credit"] = int(row["Credit"])

with open("subjects_master.csv", "r") as f:
    reader = csv.DictReader(f)
    subjects_data = {}
    for row in reader:
        row = dict(row)
        subjects_data[row["subno"]] = {
            "subname": row["subname"],
            "ltp": row["ltp"],
            "crd": int(row["crd"]),
        }

with open("names-roll.csv", "r") as f:
    reader = csv.DictReader(f)
    names_data = {}
    for row in reader:
        row = dict(row)
        names_data[row["Roll"]] = row["Name"]
        data[row["Roll"]] = {}

grade_dict = {
    "AA": 10,
    "AB": 9,
    "BB": 8,
    "BC": 7,
    "CC": 6,
    "CD": 5,
    "DD": 4,
    "DD*": 4,
    "F": 0,
    "F*": 0,
}

for row in grades_data:
    sem = row["Sem"]
    if sem not in data[row["Roll"]]:
        data[row["Roll"]][sem] = []
    data[row["Roll"]][sem].append(
        {
            "code": row["SubCode"],
            "credit": row["Credit"],
            "grade": row["Grade"],
        }
    )

for roll in names_data:
    prepare_file(data, subjects_data, names_data, grade_dict, roll)

def transcripts_in_range(range_of_roll):
    conv_upper=range_of_roll.upper()
    txt=conv_upper.split('-')
    half_roll=txt[0][0:6]
    start=int(txt[0][-2:])
    end=int(txt[1][-2:])
    roll_list=[]
    input_roll_list=[]
    for roll in names_data:
        roll_list.append(roll)
    for i in range(start,end+1):
        input_roll=half_roll+str(i)
        if input_roll in roll_list:
            input_roll_list.append(input_roll)
        else:
            print(input_roll,'Roll Number does not exit')
    location1="TranscriptsIITP"
    if not os.path.exists(location1):
        os.makedirs(location1)
    for  roll in input_roll_list:
        o=win32com.client.Dispatch("Excel.Application",pythoncom.CoInitialize())
        o.Visible = False
        wb_path = "E:\\New folder\\Work 2\\output\\"+roll+".xlsx"
        wb = o.Workbooks.Open(wb_path)
        ws_index_list = [1] #say you want to print these sheets
        path_to_pdf = "E:\\New folder\\Work 2\\TranscriptsIITP\\"+ roll+".pdf"
        for index in ws_index_list:
            ws = wb.Worksheets[index - 1]
            ws.PageSetup.Zoom = False
            ws.PageSetup.paperSize = 8
            ws.PageSetup.orientation = 2
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    wb.Close()

def transcripts_all():
    location1="Transcriptsall"
    if not os.path.exists(location1):
        os.makedirs(location1)
    for  roll in names_data:
        o=win32com.client.Dispatch("Excel.Application",pythoncom.CoInitialize())
        o.Visible = False
        wb_path = "E:\\New folder\\Work 2\\output\\"+roll+".xlsx"
        wb = o.Workbooks.Open(wb_path)
        ws_index_list = [1] #say you want to print these sheets
        path_to_pdf = "E:\\New folder\\Work 2\\Transcriptsall\\"+ roll+".pdf"
        for index in ws_index_list:
            ws = wb.Worksheets[index - 1]
            ws.PageSetup.Zoom = False
            ws.PageSetup.paperSize = 8
            ws.PageSetup.orientation = 2
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    wb.Close()
   
@app.route("/frontend", methods=["POST", "GET"])
def frontend():
    if request.method == "POST":
        nm = request.form["range"]
        range_of_roll=nm
        if request.form["submit_button"] == "Generate required Transcripts":
            transcripts_in_range(range_of_roll)
        elif request.form["submit_button"] == "Generate all Transcripts":
            transcripts_all()
        return redirect(url_for("success", name=''))
    else:
        range_of_roll = request.form["range"]
        return redirect(url_for("success", name=''))

if __name__ == "__main__":
    app.run(debug=True)
